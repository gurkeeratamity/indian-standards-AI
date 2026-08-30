"""
check_standards_coverage.py
----------------------------
Compares your master Indian Standards (IS) catalogue (an Excel workbook)
against the actual PDF files you have downloaded in data/standards/, and
tells you which standards are done, which are missing, and where there
are problems (duplicates, wrong years, unrecognized files, etc).

WHAT THIS SCRIPT DOES NOT DO (on purpose):
- It never downloads anything.
- It never modifies or deletes your PDF files.
- It never touches ChromaDB / the vector database.
- It never modifies ingest.py or any other script.
- It never overwrites your source Excel workbook, unless you deliberately
  flip the UPDATE_ORIGINAL_WORKBOOK switch below to True.
- It never invents an IS number, year, or part it could not actually find.

HOW TO RUN IT:
    python3 check_standards_coverage.py

WHAT IT PRODUCES:
    data/processed/standards_coverage_report.csv
    data/processed/standards_coverage_report.json
    data/processed/pending_standards.csv
    (optionally) a copy of your workbook with a new "PDF_Coverage" sheet

See the comments throughout for exactly how each part works.
"""

import os
import re
import sys
import json
import hashlib
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# 1. CONFIGURATION - change these if your project layout is different
# ---------------------------------------------------------------------------

# Where to look for the master catalogue workbook. The script will try each
# of these paths in order and use the first one that exists.
MASTER_XLSX_CANDIDATES = [
    "dairy.bis.xlsx",
    "Dairy.bis.xlsx",
]

# Folder containing the IS standard PDFs you have already downloaded.
# Scanned recursively, so PDFs inside subfolders are found too.
STANDARDS_DIR = Path("data/standards")

# Where reports (and the cache) get written.
OUTPUT_DIR = Path("data/processed")
CACHE_FILE = OUTPUT_DIR / "pdf_scan_cache.json"

# SAFETY SWITCH: if True, the script will add/update a "PDF_Coverage" sheet
# directly inside your original workbook. If False (the safe default), it
# writes a brand-new copy of the workbook instead, so your original file is
# never touched. Only flip this to True once you are comfortable with it.
UPDATE_ORIGINAL_WORKBOOK = False

# How many pages of a PDF to inspect when the filename alone isn't enough
# to identify its IS number.
PDF_PAGES_TO_INSPECT = 2


# ---------------------------------------------------------------------------
# 2. NORMALIZATION - turning messy text into a clean (is_number, part, year)
# ---------------------------------------------------------------------------
#
# The same standard can show up written in many different ways:
#   "IS 10484 : 2021"
#   "IS 10484:2021"
#   "IS_10484_2021.pdf"
#   "10484_2021.pdf"
#   "IS 1224 (Part 1):1977"
#   "IS_1224_Part_1_1977.pdf"
#
# normalize_is_string() turns any of these into a small dictionary:
#   {"is_number": "10484", "part": None,  "year": "2021"}
#   {"is_number": "1224",  "part": "1",   "year": "1977"}
#
# Two standards are considered "the same" only if is_number AND part match.
# Year is deliberately NOT part of that matching key - a year difference is
# instead reported as a YEAR_MISMATCH, rather than silently causing a
# non-match.

def normalize_is_string(raw):
    """Extract {is_number, part, year} from a messy IS reference string."""
    if not raw:
        return {"is_number": None, "part": None, "year": None}

    text = str(raw).upper()
    text = text.replace(".PDF", "")
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()

    # Look for "IS" followed by 2-6 digits, e.g. "IS 10484" or "IS10484".
    is_match = re.search(r"\bIS\s*0*([0-9]{2,6})\b", text)
    is_number = is_match.group(1) if is_match else None

    # Some filenames drop the "IS" prefix entirely, e.g. "10484_2021.pdf".
    # In that case, fall back to a leading run of digits.
    if is_number is None:
        num_match = re.search(r"^0*([0-9]{2,6})\b", text)
        if num_match:
            is_number = num_match.group(1)

    # Look for a part number: "PART 1", "(PART 1)", "PART1", or a bare "P1".
    part = None
    part_match = re.search(r"PART\s*0*([0-9]{1,2})", text)
    if part_match:
        part = part_match.group(1)
    else:
        p_match = re.search(r"\bP\s*0*([0-9]{1,2})\b", text)
        if p_match:
            part = p_match.group(1)

    # Look for a 4-digit year (1900-2099). Some standards are dual-listed
    # with an adopted ISO standard and its own year, e.g.
    # "IS 1479 (Part 4) : 2018/ISO 5764 : 2009" - here the FIRST year (2018)
    # is the actual IS revision year, and the second (2009) belongs to the
    # referenced ISO standard, so we deliberately take the first match.
    year = None
    year_candidates = re.findall(r"\b(19[0-9]{2}|20[0-9]{2})\b", text)
    if year_candidates:
        year = year_candidates[0]

    return {"is_number": is_number, "part": part, "year": year}


def canonical_key(is_number, part):
    """Build the matching key used to compare catalogue rows and PDFs."""
    if is_number is None:
        return None
    if part:
        return f"IS{is_number}_P{part}"
    return f"IS{is_number}"


# ---------------------------------------------------------------------------
# 3. FINDING THE IS-NUMBER REFERENCE INSIDE PDF TEXT (used as a fallback)
# ---------------------------------------------------------------------------
#
# If a PDF's filename doesn't give us a usable IS number, we open the file
# and search the first couple of pages for something that looks like an IS
# reference (e.g. "IS 10484 : 2021" printed on the cover page). We only feed
# the *matched snippet* into normalize_is_string(), not the whole page, so
# stray letters/numbers elsewhere on the page can't confuse the part/year
# extraction.

IS_REFERENCE_PATTERN = re.compile(
    r"IS\s*[0-9]{2,6}"                      # IS 10484
    r"(?:\s*\(?\s*PART\s*[0-9]{1,2}\)?)?"   # optional (Part 1)
    r"\s*:?\s*(?:19|20)[0-9]{2}",           # : 2021
    re.IGNORECASE,
)


def find_is_reference_in_pdf(pdf_path, pages_to_check=PDF_PAGES_TO_INSPECT):
    """
    Open a PDF and look for an IS-standard reference in its first pages.
    Returns (normalized_dict, page_count) on success, or (None, page_count)
    if nothing recognizable was found. Raises an exception if the PDF
    cannot be opened at all (caller treats that as NEEDS_REVIEW).
    """
    import pymupdf  # imported here so the rest of the script works even

    doc = pymupdf.open(pdf_path)
    page_count = len(doc)

    text_to_search = ""
    for i in range(min(pages_to_check, page_count)):
        text_to_search += doc[i].get_text() + "\n"
    doc.close()

    match = IS_REFERENCE_PATTERN.search(text_to_search)
    if not match:
        return None, page_count

    return normalize_is_string(match.group(0)), page_count


# ---------------------------------------------------------------------------
# 4. LOADING THE MASTER CATALOGUE FROM THE EXCEL WORKBOOK
# ---------------------------------------------------------------------------

def find_master_workbook():
    """Return the path to the master workbook, or exit with a clear error."""
    for candidate in MASTER_XLSX_CANDIDATES:
        if Path(candidate).exists():
            return candidate

    sys.exit(
        "ERROR: Could not find the master catalogue workbook.\n"
        f"Looked for: {MASTER_XLSX_CANDIDATES}\n"
        "Place the workbook in the current folder, or edit "
        "MASTER_XLSX_CANDIDATES at the top of this script."
    )


def _find_header_row_and_columns(sheet):
    """
    
    Find the catalogue header row and map the columns used by this project.
    """

    for row_idx in range(1, 11):
        column_map = {}

        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value

            if value is None:
                continue

            label = str(value).strip().lower()

            if label in ("is_number", "is number", "is no", "is", "is_no"):
                column_map["is_no"] = col_idx

            elif label == "title":
                column_map["title"] = col_idx

            elif label == "year":
                column_map["year"] = col_idx

            elif label in ("catalogue_no", "catalogue no", "catalogue number"):
                column_map["catalogue_no"] = col_idx

        if "is_no" in column_map and "title" in column_map:
            return row_idx, column_map

    return None, None


def load_master_catalogue(path):
    """
    Open the workbook, find the catalogue sheet, and return a list of
    dictionaries - one per catalogue row - with the fields we need.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    print(f"Opened workbook: {path}")
    print(f"Sheets found: {wb.sheetnames}")

    # Prefer a sheet literally named "Standards_Master" if it exists.
    sheet_names_to_try = list(wb.sheetnames)
    if "Standards_Master" in sheet_names_to_try:
        sheet_names_to_try.remove("Standards_Master")
        sheet_names_to_try.insert(0, "Standards_Master")

    chosen_sheet = None
    header_row = None
    column_map = None

    for name in sheet_names_to_try:
        sheet = wb[name]
        found_row, found_columns = _find_header_row_and_columns(sheet)
        if found_row:
            chosen_sheet = sheet
            header_row = found_row
            column_map = found_columns
            print(f"Using sheet '{name}' (header row {header_row}) "
                  f"with columns: {column_map}")
            break

    if chosen_sheet is None:
        sys.exit(
            "ERROR: Could not find a catalogue sheet with recognizable "
            "'IS No' and 'Title' columns in any sheet of the workbook.\n"
            "Rename the relevant columns, or adjust "
            "_find_header_row_and_columns() to match your headers."
        )

    catalogue_rows = []
    catalogue_counter = 0
    for row_idx in range(header_row + 1, chosen_sheet.max_row + 1):
        is_no_cell = chosen_sheet.cell(row=row_idx, column=column_map["is_no"]).value
        title_cell = chosen_sheet.cell(row=row_idx, column=column_map["title"]).value

        # Skip rows that are section headers or completely blank.
        if not is_no_cell:
            continue

        catalogue_counter += 1

        norm = normalize_is_string(is_no_cell)
        if norm["is_number"] is None:
            # We found a row but couldn't extract an IS number from it.
            # Keep it so it still appears in the report, flagged for review.
            print(f"WARNING: could not parse IS number from row {row_idx}: "
                  f"{is_no_cell!r}")

        explicit_year = None
        if "year" in column_map:
            year_cell = chosen_sheet.cell(row=row_idx, column=column_map["year"]).value
            if year_cell:
                explicit_year = str(year_cell).strip()

        catalogue_no = catalogue_counter
        if "catalogue_no" in column_map:
            raw_no = chosen_sheet.cell(row=row_idx, column=column_map["catalogue_no"]).value
            if raw_no is not None:
                catalogue_no = raw_no

        catalogue_rows.append({
            "catalogue_no": catalogue_no,
            "raw_is_text": str(is_no_cell).strip(),
            "title": str(title_cell).strip() if title_cell else "",
            "is_number": norm["is_number"],
            "part": norm["part"],
            "catalogue_year": explicit_year or norm["year"],
            "canonical_key": canonical_key(norm["is_number"], norm["part"]),
        })

    return catalogue_rows


# ---------------------------------------------------------------------------
# 5. SCANNING THE PDF FOLDER (recursively) WITH CACHING
# ---------------------------------------------------------------------------

def scan_standards_folder(root):
    """Recursively find every .pdf file under root. Returns a sorted list
    of Path objects."""
    if not root.exists():
        sys.exit(
            f"ERROR: PDF folder '{root}' does not exist.\n"
            "Create it and put your downloaded IS standard PDFs inside, "
            "or update STANDARDS_DIR at the top of this script."
        )
    return sorted(root.rglob("*.pdf"))


def load_cache():
    """Load the metadata cache from disk, or return an empty one."""
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_cache(cache):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)


def _cache_key_for_file(path, stat_result):
    """A cache entry becomes stale automatically if the file's size or
    modified time changes, so we don't need to re-hash file contents."""
    return f"{path}|{stat_result.st_size}|{int(stat_result.st_mtime)}"


def get_pdf_metadata(path, cache):
    """
    Work out {is_number, part, year, source, readable, pages} for one PDF,
    using the filename first and falling back to reading the file's first
    pages only if needed. Uses the cache to avoid re-opening unchanged
    files on repeat runs.
    """
    stat_result = path.stat()
    cache_key = _cache_key_for_file(str(path), stat_result)

    if cache_key in cache:
        return cache[cache_key]

    filename_norm = normalize_is_string(path.stem)

    result = {
        "is_number": filename_norm["is_number"],
        "part": filename_norm["part"],
        "year": filename_norm["year"],
        "source": "filename",
        "readable": True,
        "pages": None,
    }

    # If the filename didn't give us an IS number, look inside the PDF.
    if result["is_number"] is None:
        try:
            content_norm, page_count = find_is_reference_in_pdf(path)
            result["pages"] = page_count
            if content_norm and content_norm["is_number"]:
                result["is_number"] = content_norm["is_number"]
                result["part"] = content_norm["part"]
                result["year"] = content_norm["year"]
                result["source"] = "content"
        except Exception as exc:  # noqa: BLE001 - any parse failure -> review
            result["readable"] = False
            result["error"] = str(exc)

    cache[cache_key] = result
    return result


# ---------------------------------------------------------------------------
# 6. MATCHING PDFS TO CATALOGUE ROWS AND DECIDING STATUS
# ---------------------------------------------------------------------------

STATUS_COMPLETED = "COMPLETED"
STATUS_PENDING = "PENDING"
STATUS_DUPLICATE = "DUPLICATE_PDF"
STATUS_YEAR_MISMATCH = "YEAR_MISMATCH"
STATUS_PART_MISMATCH = "PART_MISMATCH"
STATUS_UNRECOGNIZED = "UNRECOGNIZED_PDF"
STATUS_NEEDS_REVIEW = "NEEDS_REVIEW"


def build_pdf_index(pdf_files, cache):
    """
    Inspect every PDF and group them by canonical key. Returns:
      - matches: dict canonical_key -> list of pdf info dicts
      - part_mismatches: list of pdf info dicts that reference a known IS
        number but an unrecognized/ambiguous part
      - unrecognized: list of pdf info dicts with no identifiable IS number
      - unreadable: list of pdf info dicts that could not be opened at all
    """
    matches = {}
    part_mismatches = []
    unrecognized = []
    unreadable = []

    for path in pdf_files:
        meta = get_pdf_metadata(path, cache)
        relative_path = str(path)

        entry = {
            "filename": path.name,
            "relative_path": relative_path,
            "is_number": meta["is_number"],
            "part": meta["part"],
            "year": meta["year"],
            "source": meta["source"],
        }

        if not meta.get("readable", True):
            entry["error"] = meta.get("error", "unreadable PDF")
            unreadable.append(entry)
            continue

        if meta["is_number"] is None:
            unrecognized.append(entry)
            continue

        key = canonical_key(meta["is_number"], meta["part"])
        matches.setdefault(key, []).append(entry)

    return matches, part_mismatches, unrecognized, unreadable


def determine_row_status(row, matches, multipart_numbers):
    """Work out the status/notes/pdf info for one catalogue row."""
    key = row["canonical_key"]

    if key is None:
        return {
            **row,
            "pdf_found": False,
            "pdf_filename": "",
            "pdf_year": "",
            "relative_pdf_path": "",
            "status": STATUS_NEEDS_REVIEW,
            "notes": "Could not parse an IS number from the catalogue row itself.",
        }

    candidates = matches.get(key, [])

    if not candidates:
        return {
            **row,
            "pdf_found": False,
            "pdf_filename": "",
            "pdf_year": "",
            "relative_pdf_path": "",
            "status": STATUS_PENDING,
            "notes": "",
        }

    if len(candidates) > 1:
        filenames = ", ".join(c["filename"] for c in candidates)
        return {
            **row,
            "pdf_found": True,
            "pdf_filename": filenames,
            "pdf_year": candidates[0]["year"] or "",
            "relative_pdf_path": ", ".join(c["relative_path"] for c in candidates),
            "status": STATUS_DUPLICATE,
            "notes": f"{len(candidates)} PDF files matched this same standard.",
        }

    match = candidates[0]

    if match["year"] is None:
        return {
            **row,
            "pdf_found": True,
            "pdf_filename": match["filename"],
            "pdf_year": "",
            "relative_pdf_path": match["relative_path"],
            "status": STATUS_NEEDS_REVIEW,
            "notes": "PDF found, but its year could not be determined "
                     "from the filename or document content.",
        }

    if row["catalogue_year"] and match["year"] != row["catalogue_year"]:
        return {
            **row,
            "pdf_found": True,
            "pdf_filename": match["filename"],
            "pdf_year": match["year"],
            "relative_pdf_path": match["relative_path"],
            "status": STATUS_YEAR_MISMATCH,
            "notes": f"Catalogue year {row['catalogue_year']} != PDF year {match['year']}.",
        }

    return {
        **row,
        "pdf_found": True,
        "pdf_filename": match["filename"],
        "pdf_year": match["year"],
        "relative_pdf_path": match["relative_path"],
        "status": STATUS_COMPLETED,
        "notes": "",
    }


def find_part_mismatches(matches, catalogue_rows, unrecognized):
    """
    Some PDFs reference a real IS number but with a part that doesn't
    exist in the catalogue (e.g. the catalogue only has Parts 1-5 for
    IS 1479, but a PDF is filed as "Part 6", or a PDF has no part at all
    for an IS number that the catalogue splits into several parts).
    These are pulled out of `unrecognized` and reported separately as
    PART_MISMATCH problems.
    """
    catalogue_keys = {row["canonical_key"] for row in catalogue_rows if row["canonical_key"]}
    is_numbers_with_parts = {
        row["is_number"] for row in catalogue_rows if row["is_number"] and row["part"]
    }

    part_mismatches = []
    still_unrecognized = []

    for entry in unrecognized:
        # This branch only ever sees entries with is_number is None already
        # filtered out upstream, so nothing to do here for those.
        still_unrecognized.append(entry)

    # Also check matched PDFs whose key isn't a real catalogue key purely
    # because of a part mismatch (e.g. filename says Part 6, catalogue only
    # goes up to Part 5). These live inside `matches` under a key that
    # doesn't correspond to any catalogue row.
    orphan_keys = [k for k in matches if k not in catalogue_keys]
    for key in orphan_keys:
        for entry in matches[key]:
            if entry["is_number"] in is_numbers_with_parts:
                part_mismatches.append(entry)
            else:
                still_unrecognized.append(entry)

    return part_mismatches, still_unrecognized


# ---------------------------------------------------------------------------
# 7. WRITING THE OUTPUT REPORTS (CSV + JSON)
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "catalogue_no", "standard_id", "is_number", "title", "catalogue_year",
    "pdf_found", "pdf_filename", "pdf_year", "status", "relative_pdf_path",
    "notes",
]


def _row_for_csv(result_row):
    return {
        "catalogue_no": result_row["catalogue_no"],
        "standard_id": result_row["canonical_key"] or "",
        "is_number": result_row["is_number"] or "",
        "title": result_row["title"],
        "catalogue_year": result_row["catalogue_year"] or "",
        "pdf_found": result_row["pdf_found"],
        "pdf_filename": result_row["pdf_filename"],
        "pdf_year": result_row["pdf_year"],
        "status": result_row["status"],
        "relative_pdf_path": result_row["relative_pdf_path"],
        "notes": result_row["notes"],
    }


def write_csv(path, rows):
    import csv
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(_row_for_csv(row))


def write_json(path, result_rows, part_mismatches, unrecognized, unreadable):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "standards": [_row_for_csv(r) for r in result_rows],
        "part_mismatches": part_mismatches,
        "unrecognized_pdfs": unrecognized,
        "unreadable_pdfs": unreadable,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# ---------------------------------------------------------------------------
# 8. OPTIONAL: WRITE A PDF_Coverage SHEET INTO THE WORKBOOK
# ---------------------------------------------------------------------------

def update_excel_with_coverage(master_path, result_rows):
    """
    If the workbook has a "Standards_Master" sheet, add/update a
    "PDF_Coverage" sheet next to it. By default this writes to a NEW copy
    of the workbook so the original is never touched - see
    UPDATE_ORIGINAL_WORKBOOK at the top of this file.
    """
    wb = openpyxl.load_workbook(master_path)

    if "Standards_Master" not in wb.sheetnames:
        print("No 'Standards_Master' sheet found - skipping the optional "
              "PDF_Coverage sheet update.")
        return None

    if "PDF_Coverage" in wb.sheetnames:
        del wb["PDF_Coverage"]
    sheet = wb.create_sheet("PDF_Coverage")

    headers = [
        "Catalogue No", "IS Number", "Title", "Expected Year",
        "PDF Found", "PDF Filename", "PDF Year", "Status", "Notes",
    ]
    sheet.append(headers)

    for row in result_rows:
        sheet.append([
            row["catalogue_no"],
            row["is_number"] or "",
            row["title"],
            row["catalogue_year"] or "",
            "Yes" if row["pdf_found"] else "No",
            row["pdf_filename"],
            row["pdf_year"],
            row["status"],
            row["notes"],
        ])

    if UPDATE_ORIGINAL_WORKBOOK:
        wb.save(master_path)
        print(f"Updated '{master_path}' in place with a PDF_Coverage sheet.")
        return master_path
    else:
        output_path = OUTPUT_DIR / (Path(master_path).stem + "_with_coverage.xlsx")
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"Wrote a NEW copy with a PDF_Coverage sheet to: {output_path}")
        print("(Your original workbook was not touched. Set "
              "UPDATE_ORIGINAL_WORKBOOK = True at the top of this script "
              "if you want the sheet added directly to the original file.)")
        return output_path


# ---------------------------------------------------------------------------
# 9. PRINTING THE TERMINAL REPORT
# ---------------------------------------------------------------------------

def print_report(result_rows, part_mismatches, unrecognized, unreadable):
    total = len(result_rows)
    completed = [r for r in result_rows if r["status"] == STATUS_COMPLETED]
    pending = [r for r in result_rows if r["status"] == STATUS_PENDING]
    duplicates = [r for r in result_rows if r["status"] == STATUS_DUPLICATE]
    year_mismatches = [r for r in result_rows if r["status"] == STATUS_YEAR_MISMATCH]
    needs_review = [r for r in result_rows if r["status"] == STATUS_NEEDS_REVIEW]

    pdf_count = sum(len(str(r["relative_pdf_path"]).split(", ")) for r in result_rows if r["pdf_found"])
    pdf_count += len(unrecognized) + len(unreadable) + len(part_mismatches)

    coverage_pct = (len(completed) / total * 100) if total else 0.0

    print("=" * 40)
    print("INDIAN STANDARDS COVERAGE REPORT")
    print("=" * 40)
    print()
    print(f"Catalogue standards: {total}")
    print(f"PDFs found: {pdf_count}")
    print(f"Standards completed: {len(completed)}")
    print(f"Pending standards: {len(pending)}")
    print(f"Duplicates: {len(duplicates)}")
    print(f"Year mismatches: {len(year_mismatches)}")
    print(f"Part mismatches: {len(part_mismatches)}")
    print(f"Unrecognized PDFs: {len(unrecognized)}")
    print(f"Needs review: {len(needs_review) + len(unreadable)}")
    print()
    print(f"Coverage: {coverage_pct:.2f}%")
    print()

    print("=" * 40)
    print("COMPLETED")
    print("=" * 40)
    for i, row in enumerate(completed, start=1):
        print(f"{i}. IS {row['is_number']}:{row['catalogue_year']}")
        print(f"   PDF: {row['pdf_filename']}")
        print()

    print("=" * 40)
    print("PENDING")
    print("=" * 40)
    for i, row in enumerate(pending, start=1):
        print(f"{i}. IS {row['is_number']}:{row['catalogue_year']}")
        print(f"   Title: {row['title']}")
        print()

    print("=" * 40)
    print("PROBLEMS")
    print("=" * 40)

    for row in year_mismatches:
        print("YEAR MISMATCH:")
        print(f"IS {row['is_number']}")
        print(f"Catalogue year: {row['catalogue_year']}")
        print(f"PDF year: {row['pdf_year']}")
        print()

    for row in duplicates:
        print("DUPLICATE:")
        print(f"IS {row['is_number']}")
        print("Files:")
        for f in row["pdf_filename"].split(", "):
            print(f"  {f}")
        print()

    for entry in part_mismatches:
        print("PART MISMATCH:")
        print(f"IS {entry['is_number']} (part: {entry['part']})")
        print(f"File: {entry['filename']}")
        print()

    for entry in unrecognized:
        print("UNRECOGNIZED PDF:")
        print(f"{entry['filename']}")
        print()

    for entry in unreadable:
        print("NEEDS REVIEW (unreadable PDF):")
        print(f"{entry['filename']} - {entry.get('error', 'could not be opened')}")
        print()

    print("=" * 40)
    print()
    print("Catalogue:", total)
    print("PDFs:", pdf_count)
    print("Completed:", len(completed))
    print("Pending:", len(pending))
    print("Problems:", len(duplicates) + len(year_mismatches) + len(part_mismatches)
          + len(unrecognized) + len(needs_review) + len(unreadable))
    print(f"Coverage: {coverage_pct:.2f}%")


# ---------------------------------------------------------------------------
# 10. MAIN
# ---------------------------------------------------------------------------

def main():
    master_path = find_master_workbook()
    catalogue_rows = load_master_catalogue(master_path)

    multipart_numbers = {
        row["is_number"] for row in catalogue_rows
        if row["is_number"] and row["part"]
    }

    pdf_files = scan_standards_folder(STANDARDS_DIR)
    cache = load_cache()
    matches, _, unrecognized_raw, unreadable = build_pdf_index(pdf_files, cache)
    save_cache(cache)

    part_mismatches, unrecognized = find_part_mismatches(
        matches, catalogue_rows, unrecognized_raw
    )

    result_rows = [
        determine_row_status(row, matches, multipart_numbers)
        for row in catalogue_rows
    ]

    write_csv(OUTPUT_DIR / "standards_coverage_report.csv", result_rows)
    write_json(
        OUTPUT_DIR / "standards_coverage_report.json",
        result_rows, part_mismatches, unrecognized, unreadable,
    )

    pending_rows = [r for r in result_rows if r["status"] == STATUS_PENDING]
    write_csv(OUTPUT_DIR / "pending_standards.csv", pending_rows)

    update_excel_with_coverage(master_path, result_rows)

    print_report(result_rows, part_mismatches, unrecognized, unreadable)


if __name__ == "__main__":
    main()